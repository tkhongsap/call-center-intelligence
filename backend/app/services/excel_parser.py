"""
Excel Parser Module

Parses Excel (.xlsx) and CSV (.csv) files for RAG ingestion.
Extracts text content with metadata including sheet names, cell positions, and headers.
"""

import csv
import io
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.services.base_parser import BaseParser, ParsedDocument

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """
    Parser for Excel (.xlsx) and CSV (.csv) files.
    
    Features:
    - Multi-sheet support for Excel files
    - Header row detection
    - Cell-level metadata
    - Configurable content formatting
    """
    
    def __init__(
        self,
        include_headers: bool = True,
        row_separator: str = "\n",
        cell_separator: str = " | ",
        include_cell_refs: bool = False,
    ):
        """
        Initialize Excel parser.
        
        Args:
            include_headers: Include header row in output
            row_separator: Separator between rows
            cell_separator: Separator between cells
            include_cell_refs: Include cell references (A1, B2) in output
        """
        self.include_headers = include_headers
        self.row_separator = row_separator
        self.cell_separator = cell_separator
        self.include_cell_refs = include_cell_refs
    
    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".csv"]
    
    def parse(self, file_bytes: bytes, filename: str) -> List[ParsedDocument]:
        """
        Parse Excel or CSV file into ParsedDocuments.
        
        For Excel: Creates one document per sheet
        For CSV: Creates one document for the entire file
        """
        self._log_parse_start(filename)
        
        ext = Path(filename).suffix.lower()
        
        try:
            if ext == ".xlsx":
                documents = self._parse_xlsx(file_bytes, filename)
            elif ext == ".csv":
                documents = self._parse_csv(file_bytes, filename)
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
            
            self._log_parse_complete(filename, len(documents))
            return documents
            
        except Exception as e:
            self._log_parse_error(filename, str(e))
            raise
    
    def _parse_xlsx(self, file_bytes: bytes, filename: str) -> List[ParsedDocument]:
        """Parse Excel .xlsx file."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel parsing. "
                "Install it with: pip install openpyxl"
            )
        
        documents = []
        file_stream = io.BytesIO(file_bytes)
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Extract content and metadata
            rows_content = []
            headers = []
            row_count = 0
            col_count = 0
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                
                row_count += 1
                col_count = max(col_count, len(row))
                
                # First non-empty row as headers
                if row_idx == 1 and self.include_headers:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                    continue
                
                # Format row content
                if self.include_cell_refs:
                    cells = [
                        f"{get_column_letter(col_idx)}{row_idx}:{cell}"
                        for col_idx, cell in enumerate(row, start=1)
                        if cell is not None
                    ]
                else:
                    cells = [str(cell) for cell in row if cell is not None]
                
                if cells:
                    rows_content.append(self.cell_separator.join(cells))
            
            # Build final content
            content_parts = []
            if headers and self.include_headers:
                content_parts.append(f"Headers: {self.cell_separator.join(headers)}")
            content_parts.extend(rows_content)
            content = self.row_separator.join(content_parts)
            
            if content.strip():
                doc = ParsedDocument.create(
                    content=content,
                    source="excel",
                    filename=filename,
                    metadata={
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "column_count": col_count,
                        "headers": headers,
                        "file_type": "xlsx",
                    },
                )
                documents.append(doc)
        
        workbook.close()
        return documents
    
    def _parse_csv(self, file_bytes: bytes, filename: str) -> List[ParsedDocument]:
        """Parse CSV file - creates one document per row."""
        documents = []
        
        # Try different encodings (including Thai encodings)
        encodings = [
            "utf-8", 
            "utf-8-sig",  # UTF-8 with BOM
            "tis-620",    # Thai Industrial Standard
            "cp874",      # Windows Thai
            "iso-8859-11", # Latin/Thai
            "latin-1", 
            "cp1252",
        ]
        
        text_content = None
        for encoding in encodings:
            try:
                text_content = file_bytes.decode(encoding)
                logger.info(f"CSV decoded successfully with encoding: {encoding}")
                break
            except (UnicodeDecodeError, LookupError):
                continue
        
        if text_content is None:
            raise ValueError("Could not decode CSV file with supported encodings")
        
        # Parse CSV
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)
        
        if not rows:
            return documents
        
        headers = rows[0] if self.include_headers else []
        data_rows = rows[1:] if self.include_headers else rows
        
        # Create one document per row (for embedding)
        for row_idx, row in enumerate(data_rows):
            if not any(cell.strip() for cell in row):
                continue  # Skip empty rows
            
            # Build row content with headers as keys (only non-empty cells)
            row_parts = []
            for i, cell in enumerate(row):
                if cell.strip():
                    if i < len(headers) and headers[i].strip():
                        row_parts.append(f"{headers[i]}: {cell}")
                    else:
                        row_parts.append(cell)
            
            row_text = self.cell_separator.join(row_parts)
            
            # Create metadata with header-value pairs (only non-empty values)
            row_metadata = {}
            for i, cell in enumerate(row):
                if not cell.strip():
                    continue  # Skip empty cells
                if i < len(headers) and headers[i].strip():
                    row_metadata[headers[i]] = cell
                else:
                    row_metadata[f"column_{i+1}"] = cell
            
            if row_text.strip():
                doc = ParsedDocument.create(
                    content=row_text,
                    source="csv",
                    filename=filename,
                    metadata=row_metadata,
                )
                documents.append(doc)
        
        return documents


# Singleton instance for convenience
excel_parser = ExcelParser()
