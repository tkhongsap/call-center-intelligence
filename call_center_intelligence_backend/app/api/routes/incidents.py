"""
Incident API Routes

Handles incident data upload and management from Excel files.
"""

import io
import uuid
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
import openpyxl
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.models.incident import Incident
from app.schemas.incident import (
    IncidentResponse,
    IncidentListResponse,
    IncidentCreate,
    IncidentUpdate,
    WordRankingResponse,
)

router = APIRouter()


# Column mapping from Thai headers to model fields
COLUMN_MAPPING = {
    "Incident": "incident_number",
    "เลขที่อ้างอิง": "reference_number",
    "วันที่รับเรื่อง": "received_date",
    "วันที่ปิดเรื่อง": "closed_date",
    "ช่องทางการติดต่อ": "contact_channel",
    "ชื่อลูกค้า": "customer_name",
    "โทรศัพท์": "phone",
    "ประเภทเรื่อง": "issue_type",
    "ประเภทเรื่องรอง 1": "issue_subtype_1",
    "ประเภทเรื่องรอง 2": "issue_subtype_2",
    "ผลิตภัณฑ์": "product",
    "กลุ่มสินค้า": "product_group",
    "โรงงานผลิต": "factory",
    "รหัสการผลิต": "production_code",
    "รายละเอียดเรื่อง": "details",
    "วิธีการแก้ปัญหา": "solution",
    "วิธีการแก้ปัญหาที่ได้จาก thaibev": "solution_from_thaibev",
    "เรื่อง": "subject",
    "อำเภอ": "district",
    "จังหวัด": "province",
    "ช่องทางการสั่งซื้อ": "order_channel",
    "สถานะ": "status",
    "ผู้รับเรื่อง": "receiver",
    "ผู้ปิดเรื่อง": "closer",
    "SLA": "sla",
}


def parse_excel_date(value) -> Optional[datetime]:
    """Parse Excel date value to datetime."""
    if value is None or value == "":
        return None
    
    if isinstance(value, datetime):
        return value
    
    if isinstance(value, str):
        # Try parsing common date formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    
    return None


def parse_excel_row(row_data: dict, upload_id: str) -> dict:
    """Parse a row from Excel into incident data."""
    incident_data = {"upload_id": upload_id}
    
    for thai_header, field_name in COLUMN_MAPPING.items():
        value = row_data.get(thai_header)
        
        # Handle date fields
        if field_name in ["received_date", "closed_date"]:
            incident_data[field_name] = parse_excel_date(value)
        # Handle other fields
        elif value is not None and value != "":
            incident_data[field_name] = str(value).strip()
        else:
            incident_data[field_name] = None
    
    return incident_data


@router.post("/upload", response_model=dict)
async def upload_incidents(
    file: UploadFile = File(...),
    analyze: bool = Query(True, description="Analyze incidents with LLM to generate summary and priority"),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload incident data from Excel file.
    
    Supports Excel files (.xlsx) with Thai language headers.
    Creates a batch upload with unique upload_id for tracking.
    Optionally analyzes incidents with LLM to generate summaries and priorities.
    """
    from app.services.incident_analysis_service import get_incident_analysis_service
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xls) are supported."
        )
    
    try:
        # Read file content
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        
        # Generate upload batch ID
        upload_id = str(uuid.uuid4())
        
        # Extract headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate required columns
        if "Incident" not in headers:
            raise HTTPException(
                status_code=400,
                detail="Missing required column 'Incident'. Please check your Excel file format."
            )
        
        # Parse rows
        incidents_to_create = []
        errors = []
        success_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if all(cell is None or cell == "" for cell in row):
                    continue
                
                # Build row data dictionary
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = value
                
                # Parse incident data
                incident_data = parse_excel_row(row_data, upload_id)
                
                # Validate incident_number exists
                if not incident_data.get("incident_number"):
                    errors.append({
                        "row": row_idx,
                        "error": "Missing incident number"
                    })
                    continue
                
                # Check for duplicates in database
                result = await db.execute(
                    select(Incident).where(
                        Incident.incident_number == incident_data["incident_number"]
                    )
                )
                existing = result.scalar_one_or_none()
                
                if existing:
                    errors.append({
                        "row": row_idx,
                        "incident_number": incident_data["incident_number"],
                        "error": "Incident number already exists"
                    })
                    continue
                
                incidents_to_create.append(Incident(**incident_data))
                success_count += 1
                
            except Exception as e:
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })
        
        # Bulk insert incidents
        if incidents_to_create:
            db.add_all(incidents_to_create)
            await db.commit()
            
            # Analyze incidents with LLM if requested
            if analyze:
                analysis_service = get_incident_analysis_service()
                
                # Prepare list of incident dicts with id and details for batch processing
                incidents_for_analysis = [
                    {
                        "id": incident.incident_number,
                        "details": incident.details
                    }
                    for incident in incidents_to_create
                    if incident.details  # Only include incidents with details
                ]
                
                # Call analyze_incidents_batch() once for all incidents
                if incidents_for_analysis:
                    try:
                        results = await analysis_service.analyze_incidents_batch(incidents_for_analysis)
                        
                        # Map results back to incident objects using incident_number
                        results_map = {r["id"]: r for r in results}
                        
                        analyzed_count = 0
                        high_medium_incidents = []  # Track HIGH and MEDIUM priority incidents for alerts
                        
                        for incident in incidents_to_create:
                            if incident.incident_number in results_map:
                                result = results_map[incident.incident_number]
                                
                                # Update incidents with summary and priority from results
                                if result["summary"]:
                                    incident.summary = result["summary"]
                                if result["priority"]:
                                    incident.priority = result["priority"]
                                    
                                    # Track HIGH and MEDIUM priority incidents for alert creation
                                    from app.models.incident import PriorityLevel
                                    if result["priority"] in [PriorityLevel.HIGH, PriorityLevel.MEDIUM]:
                                        high_medium_incidents.append(incident)
                                
                                # Count successfully analyzed incidents (those with non-None results)
                                if result["summary"] or result["priority"]:
                                    analyzed_count += 1
                        
                        # Commit analysis results
                        if analyzed_count > 0:
                            await db.commit()
                        
                        # Create alerts for HIGH and MEDIUM priority incidents
                        alerts_created = 0
                        if high_medium_incidents:
                            from app.models.alert import Alert
                            from app.models.base import AlertType, Severity, AlertStatus
                            import uuid
                            from datetime import datetime, timezone
                            
                            now = datetime.now(timezone.utc)
                            
                            for incident in high_medium_incidents:
                                try:
                                    # Map priority to severity
                                    from app.models.incident import PriorityLevel
                                    severity = Severity.high if incident.priority == PriorityLevel.HIGH else Severity.medium
                                    
                                    # Create alert
                                    alert = Alert(
                                        id=str(uuid.uuid4()),
                                        type=AlertType.threshold,
                                        severity=severity,
                                        title=f"Incident {incident.incident_number} - {incident.priority.value.upper()} Priority",
                                        description=incident.summary or incident.details[:200] if incident.details else "No description available",
                                        status=AlertStatus.active,
                                        business_unit=incident.product_group,
                                        category=incident.issue_type,
                                        channel=incident.contact_channel,
                                        created_at=now,
                                        updated_at=now
                                    )
                                    
                                    db.add(alert)
                                    alerts_created += 1
                                except Exception as e:
                                    print(f"Error creating alert for incident {incident.incident_number}: {str(e)}")
                            
                            # Commit alerts
                            if alerts_created > 0:
                                await db.commit()
                        
                        return {
                            "success": True,
                            "upload_id": upload_id,
                            "filename": file.filename,
                            "total_rows": sheet.max_row - 1,
                            "success_count": success_count,
                            "analyzed_count": analyzed_count,
                            "alerts_created": alerts_created,
                            "error_count": len(errors),
                            "errors": errors[:10] if errors else [],
                            "message": f"Successfully imported {success_count} incidents, analyzed {analyzed_count} with LLM using batch processing, created {alerts_created} alerts"
                        }
                    except Exception as e:
                        print(f"Error in batch analysis: {str(e)}")
                        # Continue without analysis if batch processing fails
                        return {
                            "success": True,
                            "upload_id": upload_id,
                            "filename": file.filename,
                            "total_rows": sheet.max_row - 1,
                            "success_count": success_count,
                            "analyzed_count": 0,
                            "error_count": len(errors),
                            "errors": errors[:10] if errors else [],
                            "message": f"Successfully imported {success_count} incidents (batch analysis failed)"
                        }
        
        return {
            "success": True,
            "upload_id": upload_id,
            "filename": file.filename,
            "total_rows": sheet.max_row - 1,  # Exclude header
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors[:10] if errors else [],  # Return first 10 errors
            "message": f"Successfully imported {success_count} incidents"
        }
        
    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(
            status_code=400,
            detail="Invalid Excel file format. Please upload a valid .xlsx file."
        )
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error processing file: {str(e)}"
        )


@router.get("/", response_model=IncidentListResponse)
async def list_incidents(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=1000, description="Items per page"),
    upload_id: Optional[str] = Query(None, description="Filter by upload batch ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    all: bool = Query(False, description="Return all incidents without pagination"),
    db: AsyncSession = Depends(get_db),
):
    """
    List incidents with pagination and filtering.
    """
    # Build query
    query = select(Incident)
    
    # Apply filters
    if upload_id:
        query = query.where(Incident.upload_id == upload_id)
    if status:
        query = query.where(Incident.status == status)
    
    # Get total count
    count_query = select(func.count()).select_from(Incident)
    if upload_id:
        count_query = count_query.where(Incident.upload_id == upload_id)
    if status:
        count_query = count_query.where(Incident.status == status)
    
    total_result = await db.execute(count_query)
    total = total_result.scalar()
    
    # Apply ordering
    query = query.order_by(Incident.id.asc())
    
    # Apply pagination (unless all=true)
    if not all:
        query = query.offset((page - 1) * page_size).limit(page_size)
    
    # Execute query
    result = await db.execute(query)
    incidents = result.scalars().all()
    
    return IncidentListResponse(
        incidents=[IncidentResponse.model_validate(inc) for inc in incidents],
        total=total,
        page=page if not all else 1,
        page_size=len(incidents) if all else page_size,
    )


@router.get("/{incident_id}", response_model=IncidentResponse)
async def get_incident(
    incident_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Get a specific incident by ID.
    """
    result = await db.execute(
        select(Incident).where(Incident.id == incident_id)
    )
    incident = result.scalar_one_or_none()
    
    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found")
    
    return IncidentResponse.model_validate(incident)


@router.get("/number/{incident_number}", response_model=IncidentResponse)
async def get_incident_by_number(
    incident_number: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Get a specific incident by incident number.
    """
    result = await db.execute(
        select(Incident).where(Incident.incident_number == incident_number)
    )
    incident = result.scalar_one_or_none()
    
    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found")
    
    return IncidentResponse.model_validate(incident)


@router.patch("/{incident_id}", response_model=IncidentResponse)
async def update_incident(
    incident_id: int,
    incident_update: IncidentUpdate,
    db: AsyncSession = Depends(get_db),
):
    """
    Update an incident.
    """
    result = await db.execute(
        select(Incident).where(Incident.id == incident_id)
    )
    incident = result.scalar_one_or_none()
    
    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found")
    
    # Update fields
    update_data = incident_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(incident, field, value)
    
    await db.commit()
    await db.refresh(incident)
    
    return IncidentResponse.model_validate(incident)


@router.delete("/{incident_id}")
async def delete_incident(
    incident_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Delete an incident.
    """
    result = await db.execute(
        select(Incident).where(Incident.id == incident_id)
    )
    incident = result.scalar_one_or_none()
    
    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found")
    
    await db.delete(incident)
    await db.commit()
    
    return {"success": True, "message": "Incident deleted successfully"}


@router.get("/stats/summary")
async def get_incident_stats(
    upload_id: Optional[str] = Query(None, description="Filter by upload batch ID"),
    db: AsyncSession = Depends(get_db),
):
    """
    Get incident statistics.
    """
    # Build base query
    query = select(Incident)
    if upload_id:
        query = query.where(Incident.upload_id == upload_id)
    
    # Total count
    total_result = await db.execute(
        select(func.count()).select_from(query.subquery())
    )
    total = total_result.scalar()
    
    # Count by status
    status_query = (
        select(Incident.status, func.count())
        .group_by(Incident.status)
    )
    if upload_id:
        status_query = status_query.where(Incident.upload_id == upload_id)
    
    status_result = await db.execute(status_query)
    status_counts = status_result.all()
    
    # Count by issue type
    issue_type_query = (
        select(Incident.issue_type, func.count())
        .where(Incident.issue_type.isnot(None))
        .group_by(Incident.issue_type)
        .order_by(func.count().desc())
        .limit(10)
    )
    if upload_id:
        issue_type_query = issue_type_query.where(Incident.upload_id == upload_id)
    
    issue_type_result = await db.execute(issue_type_query)
    issue_type_counts = issue_type_result.all()
    
    return {
        "total": total,
        "by_status": {status: count for status, count in status_counts if status},
        "top_issue_types": {issue_type: count for issue_type, count in issue_type_counts},
    }


@router.get("/words/ranking", response_model=WordRankingResponse)
async def get_word_ranking(
    top_n: int = Query(20, ge=1, le=100, description="Number of top words to return"),
    upload_id: Optional[str] = Query(None, description="Filter by upload batch ID"),
    db: AsyncSession = Depends(get_db),
):
    """
    Get top N most frequently used Thai words from incident details.
    
    Uses pythainlp for Thai word tokenization and filters out stopwords.
    Returns words ranked by frequency with count and percentage.
    """
    from app.services.thai_text_service import get_thai_text_service
    
    # Build query to get all details
    query = select(Incident.details).where(Incident.details.isnot(None))
    
    if upload_id:
        query = query.where(Incident.upload_id == upload_id)
    
    result = await db.execute(query)
    details_list = [row[0] for row in result.all() if row[0]]
    
    if not details_list:
        return WordRankingResponse(
            total_words=0,
            unique_words=0,
            top_words=[]
        )
    
    # Analyze texts using Thai text service
    try:
        service = get_thai_text_service(force_reload=True)
        analysis_result = service.get_top_words(details_list, top_n=top_n)
        
        return WordRankingResponse(
            total_words=analysis_result["total_words"],
            unique_words=analysis_result["unique_words"],
            top_words=analysis_result["top_words"]
        )
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Thai text analysis not available: {str(e)}. Please install pythainlp."
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error analyzing text: {str(e)}"
        )

